import datetime
import openpyxl
import time

workbook = openpyxl.load_workbook("data.xlsx")
sheet1 = workbook['Sheet1']
sheet2 = workbook['Sheet2']


while True:

    last_row = sheet2.max_row
    last_row += 1

    # Insert current time into first column of new row
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet2.cell(row=last_row, column=1).value = current_time

    for i in range(1, sheet1.max_column):
        #Get the data from sheet1
        sheet2.cell(row=last_row, column=i+1).value = sheet1.cell(row=1, column=i).value


    #Save the file
    workbook.save("data.xlsx")

    #Wait for 5 minutes
    time.sleep(1)




